from openpyxl import load_workbook

file_path = 'file.xlsx'
workbook = load_workbook(file_path)
sheet = workbook.active

html_content = "<table border='1' width='100%' cellpadding='3'>"

for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    if any(cell_value is not None for cell_value in row):
        if i == 1:
            html_content += "<tr style='text-align: center;'>"
        else:
            html_content += "<tr style='text-align: center;'>"
        for j, cell_value in enumerate(row):
            if isinstance(cell_value, str) and '**' in cell_value:
                team, score = cell_value.split('**')
                html_content += f"<td><b>{team.strip()}</b><br>{score.strip()}</td>"
            else:
                if j == 0:  # Обрабатываем даты
                    cell_value = cell_value.strftime('%d.%m.%Y')
                html_content += f"<td>{cell_value if cell_value is not None else ''}</td>"
        html_content += "</tr>"

html_content += "</table>"

print(html_content)