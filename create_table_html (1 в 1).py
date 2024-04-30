from openpyxl import load_workbook

file_path = 'file.xlsx'
workbook = load_workbook(file_path)
sheet = workbook.active

html_content = "<table border='1' width='100%' cellpadding='3'>"

for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    if any(cell_value is not None for cell_value in row):
        html_content += "<tr style='text-align: center;'>"
        for cell_value in row:
            if i == 1:  # Если это первая строка, добавляем теги th
                html_content += f"<th>{cell_value if cell_value is not None else ''}</th>"
            else:
                html_content += f"<td>{cell_value if cell_value is not None else ''}</td>"
        html_content += "</tr>"

html_content += "</table>"

print(html_content)

# print(f'Сохранено в файл «Таблица с расписанием.docx»')
